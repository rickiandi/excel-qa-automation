import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from collections import Counter
import re, datetime, statistics, json, shutil

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

APP_TITLE="Excel QA Automation — Data-Entry Freelance"
APP_DIR=Path.home()/".excel_qa_dataentry"
PROFILE_FILE=APP_DIR/"profiles.json"
HISTORY_FILE=APP_DIR/"audit_history.json"
APP_DIR.mkdir(exist_ok=True)

EMAIL=re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE=re.compile(r"^(?:\+?62|0)\d{8,13}$")
ORDER={"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3,"INFO":4}

DEFAULT_PROFILES={
"General Data Entry":{"rules":["missing","duplicate","types","email","phone","date","trim","number_text","formula","format","outlier"]},
"Customer Database":{"rules":["missing","duplicate_id","email","phone","date","trim","types","format"]},
"Product Catalog":{"rules":["missing","duplicate_id","number_text","negative","trim","types","format","outlier"]},
"Sales Data":{"rules":["missing","duplicate_id","date","number_text","negative","formula","types","outlier"]},
"Inventory":{"rules":["missing","duplicate_id","number_text","negative","types","trim","format","outlier"]}
}

def blank(v): return v is None or (isinstance(v,str) and not v.strip())
def norm(v): return re.sub(r"\s+"," ",str(v).strip())

def add(fs,sev,ws,cell,issue,value,desc,rec,rule):
    fs.append({"Severity":sev,"Sheet":ws.title,"Cell":cell.coordinate,"Row":cell.row,"Column":cell.column,
               "Issue":issue,"Value":"" if value is None else str(value)[:500],"Description":desc,
               "Recommendation":rec,"Rule":rule})

def audit(path, enabled=None):
    enabled=set(enabled or DEFAULT_PROFILES["General Data Entry"]["rules"])
    fs=[]
    wb=openpyxl.load_workbook(path,data_only=False,keep_vba=path.suffix.lower()==".xlsm")
    for ws in wb.worksheets:
        if ws.sheet_state!="visible" and "workbook" in enabled:
            add(fs,"INFO",ws,ws["A1"],"Hidden sheet",ws.sheet_state,"Sheet tersembunyi.",
                "Pastikan memang disengaja.","WORKBOOK_HIDDEN")
        if ws.max_row==1 and ws.max_column==1 and blank(ws["A1"].value):
            if "missing" in enabled: add(fs,"INFO",ws,ws["A1"],"Empty sheet","","Sheet kosong.",
                "Hapus bila tidak diperlukan.","SHEET_EMPTY")
            continue

        headers=[]
        for c in range(1,ws.max_column+1):
            cell=ws.cell(1,c); h=norm(cell.value) if not blank(cell.value) else ""
            headers.append(h)
            if not h and "missing" in enabled:
                add(fs,"MEDIUM",ws,cell,"Empty header","","Kolom memiliki header kosong.",
                    "Berikan nama header.","HEADER_EMPTY")
        hc=Counter(h.lower() for h in headers if h)
        if "duplicate" in enabled:
            for c,h in enumerate(headers,1):
                if h and hc[h.lower()]>1:
                    add(fs,"HIGH",ws,ws.cell(1,c),"Duplicate header",h,"Header muncul lebih dari sekali.",
                        "Gunakan header unik.","HEADER_DUPLICATE")

        rows={}; cols={c:[] for c in range(1,ws.max_column+1)}
        for r in range(2,ws.max_row+1):
            vals=[ws.cell(r,c).value for c in range(1,ws.max_column+1)]
            key=tuple(v.strip() if isinstance(v,str) else v for v in vals)
            if all(blank(v) for v in vals):
                if "missing" in enabled:
                    add(fs,"LOW",ws,ws.cell(r,1),"Blank row","","Baris kosong berada di area data.",
                        "Hapus bila tidak diperlukan.","ROW_BLANK")
            elif "duplicate" in enabled and key in rows:
                add(fs,"HIGH",ws,ws.cell(r,1),"Duplicate row","",f"Identik dengan baris {rows[key]}.",
                    "Tinjau atau hapus duplikasi.","ROW_DUPLICATE")
            else:
                rows[key]=r
            for c,v in enumerate(vals,1): cols[c].append((r,v))

        for c,items in cols.items():
            h=headers[c-1] or f"Column {c}"; non=[x for x in items if not blank(x[1])]
            if not non: continue

            if "types" in enabled:
                types=set()
                for _,v in non:
                    if isinstance(v,bool): types.add("bool")
                    elif isinstance(v,(int,float)): types.add("number")
                    elif isinstance(v,(datetime.date,datetime.datetime)): types.add("date")
                    else: types.add("text")
                if len(types)>1:
                    add(fs,"MEDIUM",ws,ws.cell(1,c),"Mixed data types",h,
                        "Tipe data campuran: "+", ".join(sorted(types)),
                        "Standarkan tipe data.","TYPE_MIXED")

            if "missing" in enabled:
                for r,v in items:
                    if blank(v):
                        add(fs,"MEDIUM",ws,ws.cell(r,c),"Missing value","",f"Cell kosong pada '{h}'.",
                            "Isi jika field wajib.","VALUE_MISSING")

            if "duplicate_id" in enabled and re.search(r"(id|kode|code|nomor|no\.?)",h,re.I):
                vc=Counter(str(v).strip().lower() if isinstance(v,str) else v for _,v in non)
                for r,v in non:
                    k=str(v).strip().lower() if isinstance(v,str) else v
                    if vc[k]>1:
                        add(fs,"HIGH",ws,ws.cell(r,c),"Duplicate identifier",v,
                            f"Identifier muncul {vc[k]} kali.","Pastikan identifier unik.","ID_UNIQUE")

            if "trim" in enabled:
                for r,v in non:
                    if isinstance(v,str) and v!=v.strip():
                        add(fs,"LOW",ws,ws.cell(r,c),"Leading/trailing spaces",v,
                            "Teks memiliki spasi di awal/akhir.","Gunakan Safe Fix.","TEXT_TRIM")

            if "email" in enabled and re.search(r"e[- ]?mail",h,re.I):
                for r,v in non:
                    if not EMAIL.match(str(v).strip()):
                        add(fs,"HIGH",ws,ws.cell(r,c),"Invalid email",v,
                            "Format email tidak valid.","Periksa alamat email.","EMAIL_VALID")

            if "phone" in enabled and re.search(r"(phone|telepon|telp|hp|mobile)",h,re.I):
                for r,v in non:
                    s=re.sub(r"[\s\-\(\)]","",str(v))
                    if not PHONE.match(s):
                        add(fs,"MEDIUM",ws,ws.cell(r,c),"Invalid phone",v,
                            "Format nomor telepon berpotensi tidak valid.","Standarkan nomor.","PHONE_VALID")

            if "date" in enabled and re.search(r"(date|tanggal|tgl)",h,re.I):
                for r,v in non:
                    if isinstance(v,(datetime.date,datetime.datetime)):
                        if not 1900<=v.year<=2100:
                            add(fs,"MEDIUM",ws,ws.cell(r,c),"Suspicious date",v,
                                "Tanggal di luar rentang 1900-2100.","Tinjau tanggal.","DATE_RANGE")
                    elif isinstance(v,str):
                        try: datetime.datetime.fromisoformat(v.strip().replace("/","-"))
                        except:
                            add(fs,"HIGH",ws,ws.cell(r,c),"Invalid date",v,
                                "Nilai tidak dikenali sebagai tanggal.","Ubah ke tanggal valid.","DATE_VALID")

            if "number_text" in enabled or "negative" in enabled:
                for r,v in non:
                    if "number_text" in enabled and re.search(r"(harga|price|amount|jumlah|qty|quantity|total|saldo|balance|nominal|nilai)",h,re.I):
                        if isinstance(v,str) and re.fullmatch(r"-?\d+([.,]\d+)?",v.strip()):
                            add(fs,"MEDIUM",ws,ws.cell(r,c),"Number stored as text",v,
                                "Angka tersimpan sebagai teks.","Gunakan Safe Fix setelah verifikasi.","NUMBER_AS_TEXT")
                    if "negative" in enabled and re.search(r"(harga|price|amount|jumlah|qty|quantity|total|saldo|balance|nominal|nilai)",h,re.I):
                        if isinstance(v,(int,float)) and not isinstance(v,bool) and v<0:
                            add(fs,"MEDIUM",ws,ws.cell(r,c),"Negative value",v,
                                "Nilai negatif pada kolom numerik.","Pastikan negatif diperbolehkan.","NEGATIVE")

            if "outlier" in enabled:
                nums=[(r,float(v)) for r,v in non if isinstance(v,(int,float)) and not isinstance(v,bool)]
                if len(nums)>=8:
                    xs=[x for _,x in nums]; q=statistics.quantiles(xs,n=4,method="inclusive")
                    q1,q3=q[0],q[2]; iqr=q3-q1; lo,hi=q1-1.5*iqr,q3+1.5*iqr
                    for r,x in nums:
                        if x<lo or x>hi:
                            add(fs,"LOW",ws,ws.cell(r,c),"Potential outlier",x,
                                f"Di luar batas IQR [{lo:g}, {hi:g}]. Bukan bukti data salah.",
                                "Tinjau berdasarkan konteks pekerjaan.","OUTLIER_IQR")

            if "format" in enabled:
                styles=Counter(ws.cell(r,c).style_id for r,_ in non)
                if len(styles)>1:
                    common=styles.most_common(1)[0][0]
                    for r,v in non:
                        if ws.cell(r,c).style_id!=common:
                            add(fs,"LOW",ws,ws.cell(r,c),"Formatting inconsistency",v,
                                "Style berbeda dari mayoritas kolom.","Tinjau format.","FORMAT")

        if "formula" in enabled:
            for row in ws.iter_rows():
                for cell in row:
                    v=cell.value
                    if isinstance(v,str) and v.startswith("="):
                        u=v.upper()
                        if "#REF!" in u:
                            add(fs,"CRITICAL",ws,cell,"Broken formula reference",v,
                                "Formula mengandung #REF!.","Perbaiki referensi.","FORMULA_REF")
                        for token in ("#DIV/0!","#VALUE!","#N/A","#NAME?","#NUM!","#NULL!"):
                            if token in u:
                                add(fs,"HIGH",ws,cell,"Formula error token",v,
                                    f"Formula mengandung {token}.","Periksa formula dan input.","FORMULA_ERROR"); break
        if "workbook" in enabled and ws.merged_cells.ranges:
            add(fs,"INFO",ws,ws["A1"],"Merged cells",len(ws.merged_cells.ranges),
                "Sheet memiliki merged cells.","Pastikan merge diperlukan.","MERGED_CELLS")
    fs.sort(key=lambda f:(ORDER[f["Severity"]],f["Sheet"],f["Row"],f["Column"]))
    return fs

def save_json(path,data):
    path.write_text(json.dumps(data,ensure_ascii=False,indent=2),encoding="utf-8")

def load_json(path,default):
    try: return json.loads(path.read_text(encoding="utf-8"))
    except: return default

def report(source,fs,out):
    wb=Workbook(); s=wb.active; s.title="Summary"; c=Counter(f["Severity"] for f in fs)
    for r,(a,b) in enumerate([("Excel QA Report",""),("Source",str(source)),
        ("Generated",datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),("Total findings",len(fs)),
        ("CRITICAL",c["CRITICAL"]),("HIGH",c["HIGH"]),("MEDIUM",c["MEDIUM"]),("LOW",c["LOW"]),("INFO",c["INFO"])],1):
        s.cell(r,1,a); s.cell(r,2,b)
    s["A1"].font=Font(bold=True,size=16)
    d=wb.create_sheet("Findings"); heads=["Severity","Sheet","Cell","Row","Column","Issue","Value","Description","Recommendation","Rule"]
    d.append(heads)
    for x in d[1]: x.font=Font(bold=True)
    for f in fs: d.append([f[h] for h in heads])
    for sh in wb.worksheets:
        sh.freeze_panes="A2"
        for col in sh.columns: sh.column_dimensions[col[0].column_letter].width=min(max(len(str(x.value or "")) for x in col)+2,60)
    wb.save(out)

def fix_safe(path,fs):
    wb=openpyxl.load_workbook(path,data_only=False,keep_vba=path.suffix.lower()==".xlsm"); n=0
    for f in fs:
        cell=wb[f["Sheet"]][f["Cell"]]
        if f["Issue"]=="Leading/trailing spaces" and isinstance(cell.value,str):
            nv=cell.value.strip()
            if nv!=cell.value: cell.value=nv; n+=1
        elif f["Issue"]=="Number stored as text":
            s=str(cell.value).strip().replace(",","")
            try: cell.value=float(s) if "." in s else int(s); n+=1
            except: pass
    out=Path(path).with_name(Path(path).stem+"_cleaned"+Path(path).suffix); wb.save(out)
    return out,n

class App(tk.Tk):
    def __init__(self):
        super().__init__(); self.title(APP_TITLE); self.geometry("1320x820"); self.minsize(1080,680)
        self.path=None; self.fs=[]; self.profiles=load_json(PROFILE_FILE,DEFAULT_PROFILES)
        self.history=load_json(HISTORY_FILE,[])
        self.profile=tk.StringVar(value=list(self.profiles)[0])

        top=tk.Frame(self,padx=16,pady=14); top.pack(fill="x")
        tk.Label(top,text=APP_TITLE,font=("Segoe UI",21,"bold")).pack(anchor="w")
        tk.Label(top,text="Automated QA for Excel data-entry freelance workflows",font=("Segoe UI",10)).pack(anchor="w",pady=(2,10))
        row=tk.Frame(top); row.pack(fill="x")
        self.filevar=tk.StringVar()
        tk.Entry(row,textvariable=self.filevar).pack(side="left",fill="x",expand=True)
        tk.Button(row,text="Browse",command=self.browse,width=11).pack(side="left",padx=5)
        tk.Label(row,text="Profile:").pack(side="left",padx=(8,3))
        self.combo=ttk.Combobox(row,textvariable=self.profile,values=list(self.profiles),state="readonly",width=22)
        self.combo.pack(side="left")
        tk.Button(row,text="ANALYZE",command=self.analyze,width=12).pack(side="left",padx=5)
        tk.Button(row,text="SAFE FIX",command=self.fix,width=11).pack(side="left")
        tk.Button(row,text="EXPORT",command=self.export,width=11).pack(side="left",padx=5)

        tools=tk.Frame(self,padx=16,pady=5); tools.pack(fill="x")
        tk.Button(tools,text="Manage Profiles",command=self.manage_profiles).pack(side="left")
        tk.Button(tools,text="Audit History",command=self.show_history).pack(side="left",padx=5)
        self.status=tk.Label(self,text="Pilih Excel dan profile, lalu Analyze.",anchor="w",padx=16,pady=8); self.status.pack(fill="x")

        cmd=tk.Frame(self,padx=16,pady=5); cmd.pack(fill="x")
        tk.Label(cmd,text="Command:").pack(side="left")
        self.command=tk.Entry(cmd); self.command.pack(side="left",fill="x",expand=True,padx=6)
        tk.Button(cmd,text="EXECUTE",command=self.execute).pack(side="left")

        f=tk.Frame(self,padx=16,pady=8); f.pack(fill="both",expand=True)
        cols=("Severity","Sheet","Cell","Issue","Value","Description","Recommendation")
        self.tree=ttk.Treeview(f,columns=cols,show="headings")
        for c,w in zip(cols,[90,130,75,190,190,430,350]): self.tree.heading(c,text=c); self.tree.column(c,width=w,anchor="w")
        vs=ttk.Scrollbar(f,orient="vertical",command=self.tree.yview); hs=ttk.Scrollbar(f,orient="horizontal",command=self.tree.xview)
        self.tree.configure(yscrollcommand=vs.set,xscrollcommand=hs.set)
        self.tree.grid(row=0,column=0,sticky="nsew"); vs.grid(row=0,column=1,sticky="ns"); hs.grid(row=1,column=0,sticky="ew")
        f.rowconfigure(0,weight=1); f.columnconfigure(0,weight=1)

    def browse(self):
        p=filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xlsm"),("All files","*.*")])
        if p: self.path=Path(p); self.filevar.set(p)

    def analyze(self):
        if not self.path: return messagebox.showwarning(APP_TITLE,"Pilih file terlebih dahulu.")
        try:
            self.status.config(text="Menganalisis..."); self.update_idletasks()
            rules=self.profiles[self.profile.get()]["rules"]; self.fs=audit(self.path,rules); self.refresh(self.fs)
            c=Counter(f["Severity"] for f in self.fs)
            score=max(0,100-(c["CRITICAL"]*15+c["HIGH"]*5+c["MEDIUM"]*2+c["LOW"]*.5))
            self.status.config(text=f"QA Score {score:.1f}/100 • {len(self.fs)} findings • Profile: {self.profile.get()} • C{c['CRITICAL']} H{c['HIGH']} M{c['MEDIUM']} L{c['LOW']}")
            self.history.insert(0,{"time":datetime.datetime.now().isoformat(timespec="seconds"),"file":str(self.path),
                                   "profile":self.profile.get(),"findings":len(self.fs),"score":round(score,1)})
            self.history=self.history[:100]; save_json(HISTORY_FILE,self.history)
        except Exception as e: messagebox.showerror("Audit failed",str(e))

    def refresh(self,fs):
        for x in self.tree.get_children(): self.tree.delete(x)
        for f in fs: self.tree.insert("", "end",values=tuple(f[k] for k in ("Severity","Sheet","Cell","Issue","Value","Description","Recommendation")))

    def execute(self):
        q=self.command.get().strip().lower()
        if not self.fs: return messagebox.showinfo(APP_TITLE,"Analyze file terlebih dahulu.")
        if "critical" in q: return self.refresh([f for f in self.fs if f["Severity"]=="CRITICAL"])
        if "high" in q: return self.refresh([f for f in self.fs if f["Severity"]=="HIGH"])
        if "medium" in q: return self.refresh([f for f in self.fs if f["Severity"]=="MEDIUM"])
        if "low" in q: return self.refresh([f for f in self.fs if f["Severity"]=="LOW"])
        keys={"duplikat":"duplicate","duplicate":"duplicate","formula":"formula","email":"email","tanggal":"date",
              "kosong":"missing","outlier":"outlier","telepon":"phone","format":"format","angka":"number","spasi":"spaces"}
        for a,b in keys.items():
            if a in q: return self.refresh([f for f in self.fs if b in f["Issue"].lower()])
        self.refresh([f for f in self.fs if q in str(f).lower()])

    def fix(self):
        if not self.fs: return messagebox.showinfo(APP_TITLE,"Analyze file terlebih dahulu.")
        try:
            out,n=fix_safe(self.path,self.fs); messagebox.showinfo("Safe Fix",f"{n} perubahan aman dibuat.\n\n{out}")
        except Exception as e: messagebox.showerror("Safe Fix failed",str(e))

    def export(self):
        if not self.path:
            return messagebox.showinfo(APP_TITLE,"Pilih file terlebih dahulu.")
        # Robust fallback: if the displayed audit result exists in history but the
        # in-memory findings list is empty (e.g. after reopening the app), rebuild
        # the current audit before exporting instead of incorrectly reporting that
        # no audit exists.
        if not self.fs:
            try:
                rules=self.profiles[self.profile.get()]["rules"]
                self.fs=audit(self.path,rules)
                self.refresh(self.fs)
            except Exception as e:
                return messagebox.showerror("Export failed",f"Tidak dapat memuat hasil audit terbaru: {e}")
        p=filedialog.asksaveasfilename(defaultextension=".xlsx",initialfile="Excel_QA_Report.xlsx",filetypes=[("Excel","*.xlsx")])
        if p:
            try: report(self.path,self.fs,p); messagebox.showinfo("Export",f"Laporan disimpan:\n{p}")
            except Exception as e: messagebox.showerror("Export failed",str(e))

    def manage_profiles(self):
        win=tk.Toplevel(self); win.title("QA Profiles"); win.geometry("560x480")
        tk.Label(win,text="Reusable QA Profiles",font=("Segoe UI",14,"bold")).pack(pady=10)
        lb=tk.Listbox(win); lb.pack(fill="both",expand=True,padx=15,pady=5)
        for n in self.profiles: lb.insert("end",n)
        def new():
            name=tk.simpledialog.askstring("Profile","Nama profile:")
            if not name: return
            self.profiles[name]={"rules":["missing","duplicate","types","email","phone","date","trim","number_text","formula","format","outlier"]}
            save_json(PROFILE_FILE,self.profiles); self.combo["values"]=list(self.profiles); self.profile.set(name); lb.insert("end",name)
        tk.Button(win,text="Create Profile",command=new).pack(pady=8)

    def show_history(self):
        win=tk.Toplevel(self); win.title("Audit History"); win.geometry("820x430")
        cols=("Time","File","Profile","Findings","Score")
        tree=ttk.Treeview(win,columns=cols,show="headings")
        for c,w in zip(cols,[160,330,150,90,80]): tree.heading(c,text=c); tree.column(c,width=w)
        for h in self.history: tree.insert("", "end",values=(h.get("time",""),h.get("file",""),h.get("profile",""),h.get("findings",""),h.get("score","")))
        tree.pack(fill="both",expand=True,padx=10,pady=10)

if __name__=="__main__":
    import tkinter.simpledialog
    App().mainloop()
